from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime



def normalizar_enteros(diccionario: dict):
    """
    Lo utilizamos para evitar falsos positivos al momento de desglosar diferencias en el proceso
    de comparación entre el JSON generado y el JSON original.

    Recorre recursivamente un dict/list y convierte a int todos los floats que son realmente enteros.

    Por ejemplo: 34567456789.0 -> 34567456789
    """
    if isinstance(diccionario, dict):
        return {k: normalizar_enteros(v) for k, v in diccionario.items()}
    elif isinstance(diccionario, list):
        return [normalizar_enteros(v) for v in diccionario]
    elif isinstance(diccionario, float) and diccionario.is_integer():
        return int(diccionario)
    else:
        return diccionario


def normalizar_fechas(obj):
    """
    Recorre recursivamente un dict/list y normaliza strings que son fechas.
    - Datetime -> 'yyyy-MM-ddTHH:mm:ss'
    - Date     -> 'yyyy-MM-dd'
    """
    if isinstance(obj, dict):
        return {k: normalizar_fechas(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [normalizar_fechas(v) for v in obj]
    elif isinstance(obj, str):
        # Si trae "T" asumimos que es Datetime
        if "T" in obj:
            try:
                dt = datetime.fromisoformat(obj)
                return dt.strftime("%Y-%m-%dT%H:%M:%S")
            except ValueError:
                return obj
        else:
            try:
                dt = datetime.strptime(obj, "%Y-%m-%d")
                return dt.strftime("%Y-%m-%d")
            except ValueError:
                return obj
    else:
        return obj


class JSONsComparator:
    """
    Clase encargada de comparar dos JSON y anexar sus diferencias en un archivo XLSX.
    """

    @staticmethod
    def comparar_json_a_excel(original: dict, generado: dict, ruta_salida: str, list_key_map: dict | None = None):
        """
        Compara dos JSON a nivel de hoja (clave-valor) y exporta diferencias a Excel.
        Permite desglosar campo a campo en la salida de diferencias.

        Params:
            original: JSON original.
            generado: JSON generado.
            ruta_salida: path del .xlsx de salida.
            list_key_map: mapeo opcional {ruta_lista: 'clave_del_item'} para alinear listas de dicts por clave.
                          Ej: {'AdminDeclaracion/0/Concepto/IngresosRCPF/DetIngresosMesesAnt/DetIngresosMesesAnteriores': 'mes'}
        """
        if list_key_map is None:
            list_key_map = {}

        def is_scalar(x):
            return not isinstance(x, (dict, list))

        def flatten(obj, base_path=""):
            """
            Esta función permite desdoblar los diccionarios por completo hasta cada una de sus hojas.

            Devuelve {ruta: valor_esc}
            - Dicts: ruta/clave
            - Listas: ruta/idx por defecto; si base_path está en list_key_map y los items son dicts con esa clave,
                      usa ruta/valorClave en vez de índice.
            """
            out = {}

            if is_scalar(obj):
                out[base_path or "/"] = obj
                return out

            if isinstance(obj, dict):
                for k, v in obj.items():
                    new_path = f"{base_path}/{k}" if base_path else str(k)
                    out.update(flatten(v, new_path))
                return out

            # Es una lista
            key_field = list_key_map.get(base_path)
            if key_field and all(isinstance(it, dict) and (key_field in it) for it in obj):
                # Alinear por la clave lógica
                for it in obj:
                    key_val = str(it[key_field])
                    new_path = f"{base_path}/{key_val}"
                    out.update(flatten(it, new_path))
            else:
                # Alinear por índice -> Caso por default (caso que aplica)
                for idx, it in enumerate(obj):
                    new_path = f"{base_path}/{idx}" if base_path else str(idx)
                    out.update(flatten(it, new_path))

            return out

        flat_o = flatten(original)
        flat_g = flatten(generado)

        # Unir todas las rutas posibles
        all_paths = sorted(set(flat_o.keys()) | set(flat_g.keys()))

        # Crear Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Diferencias"

        # -------------- EN ESTA PARTE SE ESPECIFICAN CARACTERÍSTICAS DE FORMATO --------------

        # Estilo de cabecera
        header_font = Font(bold=True, color="FFFFFF", size=12)  # Negrita, blanco
        header_fill = PatternFill("solid", fgColor="4F81BD")  # Fondo azul
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Ajustar alto de la cabecera (fila 1)
        ws.row_dimensions[1].height = 22

        # Insertar cabecera con estilo
        headers = ["Ruta", "Valor JSON Original", "Valor JSON Generado"]
        ws.append(headers)
        for col, title in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Ajustar ancho de columnas
        ws.column_dimensions["A"].width = 60
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 40

        # Estilo general para celdas
        default_font = Font(name="Calibri", size=11)
        wrap_alignment = Alignment(wrap_text=True, vertical="center")

        # Bordes para diferenciar
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        # -------------- EN ESTA PARTE SE DESGLOSAN LAS DIFERENCIAS --------------
        diffs = 0
        for path in all_paths:
            v1 = flat_o.get(path, None)
            v2 = flat_g.get(path, None)

            if v1 != v2 or type(v1) != type(v2):
                ws.append([path, "" if v1 is None else str(v1), "" if v2 is None else str(v2)])
                diffs += 1

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=3):
            for cell in row:
                cell.font = default_font
                cell.alignment = wrap_alignment
                cell.border = thin_border

        wb.save(ruta_salida)

        if diffs == 0:
            print("No hay diferencias encontradas. Archivos JSON idénticos")
        else:
            print(f"Diferencias guardadas en {ruta_salida} (Total: {diffs})")

        print("------------------------------------------------------")